from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook


def update_status(
    path: str | Path,
    test_case_ids: list[str],
    status: str,
    sheet_name: str | None = None,
    id_column: str = "id",
    status_column: str = "status",
) -> int:
    workbook_path = Path(path)
    wb = load_workbook(workbook_path)
    ws = wb[sheet_name] if sheet_name else wb.active

    headers = [str(cell.value or "").strip().lower() for cell in ws[1]]
    if id_column.lower() not in headers:
        raise ValueError(f"ID column not found: {id_column}")

    id_idx = headers.index(id_column.lower()) + 1
    if status_column.lower() in headers:
        status_idx = headers.index(status_column.lower()) + 1
    else:
        status_idx = len(headers) + 1
        ws.cell(row=1, column=status_idx).value = status_column

    wanted = set(test_case_ids)
    updated = 0
    for row in range(2, ws.max_row + 1):
        value = str(ws.cell(row=row, column=id_idx).value or "")
        if value in wanted:
            ws.cell(row=row, column=status_idx).value = status
            updated += 1

    wb.save(workbook_path)
    return updated


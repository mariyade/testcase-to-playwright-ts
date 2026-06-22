from __future__ import annotations

from pathlib import Path
from typing import Any
from zipfile import ZipFile
import xml.etree.ElementTree as ET

try:
    from openpyxl import load_workbook
except ModuleNotFoundError:
    load_workbook = None

from agent.models import InputSource, Priority, TestCase, TestSpec, TestType
from agent.stage1_ticket_parser.readers.parser_utils import (
    enum_or_default,
    first,
    normalise_header,
    split_lines,
    split_tags,
)


def read_excel_spec(
    path: str | Path,
    sheet_name: str | None = None,
    status_filter: str | None = None,
    limit: int | None = None,
) -> TestSpec:
    workbook_path = Path(path)
    rows = _read_rows(workbook_path, sheet_name)
    if not rows:
        raise ValueError(f"No rows found in {workbook_path}")

    headers = [normalise_header(value) for value in rows[0]]
    test_cases: list[TestCase] = []
    affected_pages: set[str] = set()
    raw_lines: list[str] = []
    normalized_status = _normalize_status(status_filter)

    for index, row in enumerate(rows[1:], start=2):
        data: dict[str, Any] = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
        if not any(data.values()):
            continue
        if normalized_status and _normalize_status(first(data, "automation status", "status", default="")) != normalized_status:
            continue

        tc_id = str(first(data, "id", "test case id", "tc id", default=f"TC_{index:03d}"))
        title = str(first(data, "title", "test case", "scenario", default=f"Test case {index}"))
        steps = split_lines(first(data, "steps", "test steps", "action", default=""))
        expected = str(first(data, "expected", "expected result", "result", default=""))
        page = str(first(data, "page", "affected page", "module", default="")).strip()

        if page:
            affected_pages.add(page)

        test_cases.append(
            TestCase(
                id=tc_id,
                title=title,
                priority=enum_or_default(Priority, first(data, "priority", default="High"), Priority.HIGH),
                type=enum_or_default(TestType, first(data, "type", "test type", default="Regression"), TestType.REGRESSION),
                preconditions=split_lines(first(data, "preconditions", "precondition", default="")),
                steps=steps,
                expected_result=expected,
                tags=_tags(data),
            )
        )
        raw_lines.append(str(data))
        if limit and len(test_cases) >= limit:
            break

    return TestSpec(
        source=InputSource.EXCEL,
        source_id=str(workbook_path),
        title=workbook_path.stem,
        affected_pages=sorted(affected_pages),
        test_cases=test_cases,
        raw_content="\n".join(raw_lines),
    )


def _read_rows(workbook_path: Path, sheet_name: str | None) -> list[list[Any]]:
    if load_workbook:
        wb = load_workbook(workbook_path)
        ws = wb[sheet_name] if sheet_name else wb.active
        return [list(row) for row in ws.iter_rows(values_only=True)]
    return _read_rows_from_xlsx_xml(workbook_path, sheet_name)


def _read_rows_from_xlsx_xml(workbook_path: Path, sheet_name: str | None) -> list[list[str]]:
    ns = {
        "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
        "rel": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
        "pkgrel": "http://schemas.openxmlformats.org/package/2006/relationships",
    }
    with ZipFile(workbook_path) as archive:
        worksheet_path = _worksheet_path(archive, sheet_name, ns)
        root = ET.fromstring(archive.read(worksheet_path))
        shared_strings = _shared_strings(archive, ns)
        rows: list[list[str]] = []
        for row in root.findall(".//main:sheetData/main:row", ns):
            values: list[str] = []
            for cell in row.findall("main:c", ns):
                index = _column_index(cell.attrib["r"])
                while len(values) <= index:
                    values.append("")
                values[index] = _cell_value(cell, shared_strings, ns)
            rows.append(values)
        return rows


def _worksheet_path(archive: ZipFile, sheet_name: str | None, ns: dict[str, str]) -> str:
    workbook = ET.fromstring(archive.read("xl/workbook.xml"))
    rels = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
    rel_map = {rel.attrib["Id"]: _resolve_workbook_target(rel.attrib["Target"]) for rel in rels}
    sheets = workbook.find("main:sheets", ns)
    if sheets is None:
        raise ValueError("Workbook has no sheets")

    selected = None
    for sheet in sheets:
        if sheet_name is None or sheet.attrib["name"] == sheet_name:
            selected = sheet
            break
    if selected is None:
        raise ValueError(f"Sheet not found: {sheet_name}")

    rel_id = selected.attrib["{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"]
    return rel_map[rel_id]


def _resolve_workbook_target(target: str) -> str:
    normalized = target.lstrip("/")
    if normalized.startswith("xl/"):
        return normalized
    return f"xl/{normalized}"


def _shared_strings(archive: ZipFile, ns: dict[str, str]) -> list[str]:
    if "xl/sharedStrings.xml" not in archive.namelist():
        return []
    root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
    values: list[str] = []
    for item in root.findall("main:si", ns):
        parts = [node.text or "" for node in item.findall(".//main:t", ns)]
        values.append("".join(parts))
    return values


def _column_index(cell_ref: str) -> int:
    letters = "".join(char for char in cell_ref if char.isalpha())
    index = 0
    for char in letters:
        index = index * 26 + ord(char.upper()) - 64
    return index - 1


def _cell_value(cell: ET.Element, shared_strings: list[str], ns: dict[str, str]) -> str:
    cell_type = cell.attrib.get("t")
    if cell_type == "inlineStr":
        node = cell.find("main:is/main:t", ns)
        return node.text if node is not None else ""
    node = cell.find("main:v", ns)
    value = node.text if node is not None else ""
    if cell_type == "s" and value:
        return shared_strings[int(value)]
    return value


def _tags(data: dict[str, Any]) -> list[str]:
    tags = split_tags(first(data, "tags", default=""))
    status = first(data, "automation status", "status", default="")
    if status:
        tags.append(f"automation_status:{_normalize_status(status)}")
    return tags


def _normalize_status(value: Any) -> str:
    return str(value or "").strip().lower().replace(" ", "_").replace("-", "_")
